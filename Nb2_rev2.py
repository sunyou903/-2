# -*- coding: utf-8 -*-
import pandas as pd
from rapidfuzz import fuzz
from pathlib import Path
import argparse
import os

SHEET = "단가대비표"
THRESHOLD = 30  # 30%

LABELS = {
    "품명": "품 명",
    "규격": "규 격",
    "단위": "단위",
    "재료비적용단가": "재료비 적용단가",
    "노무비": "노 무 비",
    "경비적용단가": "경비 적용단가",
}
# 분야 라벨 추출 (파일명에서 키워드 탐지)
def detect_label(pathlike: str) -> str:
    # 순서 중요: "건축설비"가 "건축"보다 먼저 매칭되도록
    KEYS = ["건축설비", "토목", "조경", "건축", "기계", "전기"]
    name = Path(pathlike).stem.replace(" ", "")
    for key in KEYS:
        if key in name:
            return key
    # 키워드 없으면 기본값
    return "왼쪽"  # 기본값(왼쪽/오른쪽은 main에서 덮어씁니다)

# 단가대비표 시트 로드
def load_sheet(fp, sheet=SHEET):
    return pd.read_excel(fp, sheet_name=sheet)

# 머리글 위치 탐지
def find_header_positions_multirow(df, targets, scan_rows=6):
    found = {t: [] for t in targets.values()}
    for r in range(min(scan_rows, len(df))):
        row = df.iloc[r]
        for col in df.columns:
            val = row[col]
            if isinstance(val, str):
                key = val.replace(" ", "")
                for t in targets.values():
                    if key == t.replace(" ", ""):
                        found[t].append((r, col))
    return found

# 핵심 컬럼 추출
def extract_core(df):
    pos = find_header_positions_multirow(df, LABELS, scan_rows=6)

    # 가장 많은 머리글이 잡힌 행을 헤더로 설정
    row_count = {}
    for hits in pos.values():
        for r, _c in hits:
            row_count[r] = row_count.get(r, 0) + 1
    header_row = max(row_count.items(), key=lambda x: x[1])[0]

    def pick_col(label):
        hits = [(r, c) for (r, c) in pos[label] if r == header_row]
        return hits[0][1] if hits else None

    col_map = {
        "품명": pick_col(LABELS["품명"]),
        "규격": pick_col(LABELS["규격"]),
        "단위": pick_col(LABELS["단위"]),
        "재료비적용단가": pick_col(LABELS["재료비적용단가"]),
        "노무비": pick_col(LABELS["노무비"]),
        "경비적용단가": pick_col(LABELS["경비적용단가"]),
    }

    dat = df.loc[header_row + 1:, list(col_map.values())].copy()
    dat.columns = ["품명", "규격", "단위", "재료비적용단가", "노무비", "경비적용단가"]

    # 숫자형 변환
    for col in ["재료비적용단가", "노무비", "경비적용단가"]:
        dat[col] = pd.to_numeric(dat[col], errors="coerce")

    # 품명 NaN 제거, 규격 NaN → 빈 문자열
    dat = dat.dropna(subset=["품명"]).copy()
    dat["규격"] = dat["규격"].fillna("")
    return dat

# 문자열 정규화
def norm_name(s):
    return str(s).strip()

# 규격 유사도 계산
def sim(a, b):
    if (a == "" or pd.isna(a)) and (b == "" or pd.isna(b)):
        return 100.0
    return float(fuzz.token_sort_ratio(str(a), str(b)))

# 매칭
def match_and_compare(left_df, right_df, th=THRESHOLD, left_prefix="기계", right_prefix="토목"):

    right_df = right_df.copy()
    right_df["품명_norm"] = right_df["품명"].apply(norm_name)

    rows = []
    for _, a in left_df.iterrows():
        a_name = norm_name(a["품명"])
        a_spec = a["규격"]

        # 1단계: 품명 완전일치
        exact_candidates = right_df[right_df["품명_norm"] == a_name]
        chosen = None
        match_type = None
        name_sim = spec_sim = total_sim = None

        if len(exact_candidates) > 0:
            best_idx = None
            best_spec = -1
            for idx, c in exact_candidates.iterrows():
                s = sim(a_spec, c["규격"])
                if s > best_spec:
                    best_spec, best_idx = s, idx
            c = exact_candidates.loc[best_idx]
            name_sim, spec_sim = 100.0, best_spec
            total_sim = 0.8 * name_sim + 0.2 * spec_sim
            chosen = c
            match_type = "품명완전일치"
        else:
            # 2단계: 가중 유사도 매칭
            best_idx, best_score, best_name, best_spec = None, -1, -1, -1
            for idx, c in right_df.iterrows():
                n = float(fuzz.token_sort_ratio(a_name, c["품명_norm"]))
                s = sim(a_spec, c["규격"])
                score = 0.8 * n + 0.2 * s
                if score > best_score:
                    best_score, best_idx, best_name, best_spec = score, idx, n, s

            if best_score >= th and best_idx is not None:
                c = right_df.loc[best_idx]
                chosen = c
                match_type = "가중유사도"
                name_sim, spec_sim, total_sim = best_name, best_spec, best_score

        if chosen is None:
            continue

        rows.append({
            f"{left_prefix}_품명": a["품명"],
            f"{left_prefix}_규격": a_spec,
            f"{right_prefix}_품명": chosen["품명"],
            f"{right_prefix}_규격": chosen["규격"],
            "매칭유형": match_type,
            "종합유사도(%)": round(total_sim, 1),
            "품명유사(%)": round(name_sim, 1),
            "규격유사(%)": round(spec_sim, 1),
            f"{left_prefix}_재료비적용단가": a["재료비적용단가"],
            f"{right_prefix}_재료비적용단가": chosen["재료비적용단가"],
            f"{left_prefix}_노무비": a["노무비"],
            f"{right_prefix}_노무비": chosen["노무비"],
            f"{left_prefix}_경비적용단가": a["경비적용단가"],
            f"{right_prefix}_경비적용단가": chosen["경비적용단가"],
        })
    out = pd.DataFrame(rows).sort_values(by="종합유사도(%)", ascending=False).reset_index(drop=True)
    return out

def main():
    parser = argparse.ArgumentParser(description="단가 매칭 스크립트")
    parser.add_argument("mech", help="기계 엑셀 파일 경로")
    parser.add_argument("civil", help="토목 엑셀 파일 경로")
    args = parser.parse_args()
    
    # 라벨 결정 (파일명에서 '토목/조경/건축/건축설비/기계/전기' 추출)
    left_label  = detect_label(args.mech)   or "왼쪽"
    right_label = detect_label(args.civil)  or "오른쪽"

    # 입력 파일명에서 확장자 제거
    mech_name = Path(args.mech).stem
    civil_name = Path(args.civil).stem
    # 표시용은 라벨(짧게), 내부 로그용으로는 원래 stem도 남겨둠
    OUT = Path(f"output/{left_label}_vs_{right_label}.xlsx")

    # 출력 경로 = output/파일1_vs_파일2.xlsx
    OUT = Path(f"단가대비표매칭결과/{mech_name}_vs_{civil_name}.xlsx")
    OUT.parent.mkdir(exist_ok=True)

    mech_raw = load_sheet(args.mech)
    civil_raw = load_sheet(args.civil)

    mech_core = extract_core(mech_raw)
    civil_core = extract_core(civil_raw)

    result = match_and_compare(mech_core, civil_core, THRESHOLD,
                           left_prefix=left_label, right_prefix=right_label)

    # 결과 저장 (엑셀 파일)
    result.to_excel(OUT, index=False, engine="openpyxl")


    # sheet_name을 라벨 기반으로 지정
    sheet_name = f"{left_label}_vs_{right_label}"

    # 결과 저장 (엑셀 파일, 시트명 = "토목_vs_조경" 같은 형식)
    result.to_excel(OUT, index=False, engine="openpyxl", sheet_name=sheet_name)


    # 저장된 파일 다시 열어서 마지막 열에 수식 추가
    from openpyxl import load_workbook
    wb = load_workbook(OUT)
    ws = wb[sheet_name] 

    last_col = ws.max_column + 1
    ws.cell(row=1, column=last_col, value="비교결과")

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=last_col,
                value=f"=IF(AND(J{r}=I{r}, K{r}=L{r}, M{r}=N{r}), TRUE, FALSE)")

    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT)

    print("저장 완료:", OUT)

    

if __name__ == "__main__":
    main()


# === (추가) 간이 웹서버 모드 ===
# 사용:  python Nb2_rev2.py --serve
# 또는:  WEB=1 python Nb2_rev2.py
try:
    import io, tempfile, shutil, json, os
    from flask import Flask, request, send_from_directory, jsonify
except Exception:
    pass

def _run_nb2_in_memory(a_stream, b_stream, workdir):
    os.makedirs(workdir, exist_ok=True)
    a_path = os.path.join(workdir, "A.xlsx")
    b_path = os.path.join(workdir, "B.xlsx")
    with open(a_path, "wb") as f: f.write(a_stream.read())
    with open(b_path, "wb") as f: f.write(b_stream.read())

    # 기존 main() 흐름을 그대로 재사용하기 위해 argparse 대신 직접 호출
    # load_sheet/extract_core/match_and_compare 로직 사용
    mech_raw  = load_sheet(a_path)
    civil_raw = load_sheet(b_path)
    mech_core  = extract_core(mech_raw)
    civil_core = extract_core(civil_raw)

    left_label  = detect_label(a_path) or "왼쪽"
    right_label = detect_label(b_path) or "오른쪽"
    from pathlib import Path
    mech_name = Path(a_path).stem
    civil_name = Path(b_path).stem
    out_dir = os.path.join(workdir, "단가대비표매칭결과")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{mech_name}_vs_{civil_name}.xlsx")

    result = match_and_compare(mech_core, civil_core, THRESHOLD,
                               left_prefix=left_label, right_prefix=right_label)

    # 저장 + 시트명/필터/수식 동일 적용
    result.to_excel(out_path, index=False, engine="openpyxl", sheet_name=f"{left_label}_vs_{right_label}")
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    sheet_name = f"{left_label}_vs_{right_label}"
    ws = wb[sheet_name]
    last_col = ws.max_column + 1
    ws.cell(row=1, column=last_col, value="비교결과")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=last_col, value=f"=IF(AND(J{r}=I{r}, K{r}=L{r}, M{r}=N{r}), TRUE, FALSE)")
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)

    log = f"저장 완료: {out_path}"
    return log, out_path

def _create_app_for_nb2():
    app = Flask(__name__)
    DL_DIR = tempfile.mkdtemp(prefix="nb2_out_")

    @app.post("/api/nb2")
    def api_nb2():
        a = request.files.get("mech")
        b = request.files.get("civil")
        if not a or not b:
            return jsonify({"ok": False, "log": "두 파일(mech, civil)을 모두 업로드하세요."}), 400
        work = tempfile.mkdtemp(prefix="nb2_work_")
        log, out_path = _run_nb2_in_memory(a.stream, b.stream, work)
        dst = os.path.join(DL_DIR, os.path.basename(out_path))
        shutil.copy2(out_path, dst)
        files = [{"name": os.path.basename(dst), "url": f"/download/{os.path.basename(dst)}"}]
        return jsonify({"ok": True, "log": log, "files": files})

    @app.get("/download/<path:fname>")
    def download_file(fname):
        return send_from_directory(DL_DIR, fname, as_attachment=True)

    return app

if __name__ == "__main__":
    import os, sys
    if "--serve" in sys.argv or os.environ.get("WEB") == "1":
        app = _create_app_for_nb2()
        app.run(host="0.0.0.0", port=8001, debug=False)
