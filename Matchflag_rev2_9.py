#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse, os, re, sys
import pandas as pd
from openpyxl import load_workbook

def safe_to_excel(df, path, index=False):
    import os, time
    base, ext = os.path.splitext(path); k=1
    while True:
        try:
            df.to_excel(path, index=index)
            return path
        except PermissionError:
            path = f"{base}({k}){ext}"
            k += 1
            time.sleep(0.2)

def _norm_label_s(s):
    if s is None: return None
    return str(s).replace(" ", "").replace("\u3000","").strip().lower()

def _sum_header_relaxed(ws):
    """
    공종별집계표 헤더를 느슨하게 찾는다.
    - 품명, 규격
    - 재료비 단가/노무비 단가/경비 단가 (공백 유무 허용)
    """
    needed = {
        "품명": ["품명","공종명","항목명"],
        "규격": ["규격","규 격"],
        "재료비 단가": ["재료비 단가","재료비단가","재 료 비 단 가"],
        "노무비 단가": ["노무비 단가","노무비단가","노 무 비 단 가"],
        "경비 단가": ["경비 단가","경비단가","경 비 단 가"],
    }
    # 첫 10행에서 탐색
    for r in range(1, min(ws.max_row, 10)+1):
        pos = {}
        # 열 스캔
        for c in range(1, ws.max_column+1):
            v = _norm_label_s(ws.cell(r, c).value)
            if not v: 
                continue
            for canon, alts in needed.items():
                for a in alts:
                    if _norm_label_s(a) == v:
                        if canon not in pos:
                            pos[canon] = c
                        break
        if all(k in pos for k in needed.keys()):
            return r, pos
    raise RuntimeError(f"헤더 라벨(완화) 탐색 실패: {ws.title}")


# ------------ Common helpers ------------
def _norm_commas_ws(s):
    if s is None:
        return ""
    t = str(s).replace(",", " ")
    t = " ".join(t.split())
    return t.strip()

def norm_key(key):
    if key is None:
        return ""
    s = str(key)
    if "|" in s:
        a, b = s.split("|", 1)
        return f"{_norm_commas_ws(a)}|{_norm_commas_ws(b)}"
    return _norm_commas_ws(s)

def norm(s):
    return None if s is None else str(s).strip()

def norm_label(s):
    if s is None:
        return None
    return str(s).replace(" ", "").replace("\u3000","").strip()

def find_cols(ws, required_labels, max_scan_rows=40, max_scan_cols=100):
    req_norm = {norm_label(x) for x in required_labels}
    for r in range(1, max_scan_rows + 1):
        found = {}
        for c in range(1, max_scan_cols + 1):
            nv = norm_label(ws.cell(r, c).value)
            if nv in req_norm:
                for lab in required_labels:
                    if norm_label(lab) == nv:
                        found[lab] = c
                        break
        if set(required_labels).issubset(found.keys()):
            return r, found
    raise RuntimeError(f"헤더 라벨 탐색 실패: {ws.title} -> {required_labels}")

def build_key_map(ws, header_row, pos_map):
    key_by_row = {}
    for r in range(header_row + 1, ws.max_row + 1):
        p = norm(ws.cell(r, pos_map['품명']).value)
        g = norm(ws.cell(r, pos_map['규격']).value)
        if p is None and g is None:
            continue
        key_by_row[r] = f"{p}|{g}"
    return key_by_row

# ------------ A: 일치/불일치 모두 기록 ------------
REF_RE = re.compile(r"(?:'?)((?:단가대비표)|(?:일위대가목록))(?:'?)!\$?([A-Z]{1,3})\$?(\d+)", re.I)
def run_check_A(wb):
    ul = wb['일위대가']
    up = wb['단가대비표']
    lst = wb['일위대가목록']

    ul_hr, ul_pos = find_cols(ul, {'품명','규격','단위','수량'})
    up_hr, up_pos = find_cols(up, {'품명','규격','단위'})
    ls_hr, ls_pos = find_cols(lst, {'품명','규격'})

    def is_sum_row_ul(r):
        pname = norm(ul.cell(r, ul_pos['품명']).value)
        if not pname: return False
        t = str(pname)
        return ('합' in t and '계' in t) and ('[' in t and ']' in t)

    ul_key = build_key_map(ul, ul_hr, ul_pos)
    up_key = build_key_map(up, up_hr, up_pos)
    ls_key = build_key_map(lst, ls_hr, ls_pos)

    records = []
    checked = 0

    for r in range(ul_hr + 1, ul.max_row + 1):
        cur_key = ul_key.get(r)
        if not cur_key or is_sum_row_ul(r):
            continue

        row_has_ref = False

        for c in range(1, ul.max_column + 1):
            cell = ul.cell(r, c)

            # --- 기존 참조 검사 ---
            if cell.data_type == 'f' and isinstance(cell.value, str) \
               and (('단가대비표' in cell.value) or ('일위대가목록' in cell.value)):

                for sheet_name, col, rownum in REF_RE.findall(cell.value):
                    rr = int(rownum)
                    if sheet_name.startswith('단가대비표'):
                        ref_key = up_key.get(rr)
                    else:
                        ref_key = ls_key.get(rr)
                    if not ref_key:
                        continue
                    checked += 1
                    match_status = "일치" if norm_key(ref_key) == norm_key(cur_key) else "불일치"

                    # 규격,품명에 % 포함 시 A-불일치에서 제외 처리

                    try:
                        pname_cur = norm(ul.cell(r, ul_pos["품명"]).value)
                        gname_cur = norm(ul.cell(r, ul_pos["규격"]).value)

                        if match_status == "불일치" and (
                            (pname_cur and "%" in str(pname_cur)) or
                            (gname_cur and "%" in str(gname_cur))
                        ):
                            match_status = "제외"
                    except Exception:

                        pass

                    records.append({
                        "일위대가_행": r,
                        "일위대가_품명|규격": cur_key,
                        "참조시트": sheet_name,
                        "참조셀": f"{sheet_name}!{col}{rr}",
                        "참조_품명|규격": ref_key,
                        "수식_셀": cell.coordinate,
                        "수식_일부": cell.value[:140] + ('...' if len(cell.value) > 140 else ''),
                        "일치여부": match_status
                    })
                    row_has_ref = True

        # --- 추가: 참조가 전혀 없는 경우 → 값 직접입력 여부 검사 ---
        if not row_has_ref:
            for target_col in ("수량",):
                cidx = ul_pos[target_col]
                val = ul.cell(r, cidx).value
                if val not in (None, "", 0):
                    # 직접입력: 품명에 % 포함 시 제외 처리

                    status_di = "불일치"

                    try:

                    
                        pname_cur = norm(ul.cell(r, ul_pos["품명"]).value)
                        gname_cur = norm(ul.cell(r, ul_pos["규격"]).value)

                        if (pname_cur and "%" in str(pname_cur)) or (gname_cur and "%" in str(gname_cur)):
                            status_di = "제외"
                    except Exception:

                        pass

                    records.append({

                        "일위대가_행": r,

                        "일위대가_품명|규격": cur_key,

                        "참조시트": "",

                        "참조셀": "",

                        "참조_품명|규격": "",

                        "수식_셀": ul.cell(r, cidx).coordinate,

                        "수식_일부": str(val),

                        "일치여부": status_di

                    })

    df = pd.DataFrame(records)
    summary = {
        "A_검사한_참조": checked,
        "A_일치": int((df["일치여부"]=="일치").sum()) if not df.empty else 0,
        "A_불일치": int((df["일치여부"]=="불일치").sum()) if not df.empty else 0
    }
    return summary, df

# ------------ B: 유지 (원본 로직) ------------
UL_CELLREF_RE = re.compile(r"(?:'?)일위대가(?:'?)!\$?([A-Z]{1,3})\$?(\d+)", re.I)

def run_check_B(wb):
    """
    B: 일위대가목록 ↔ 일위대가 매핑.
    - 참조된 '일위대가' 셀들에서 최근접 헤더행을 찾아 (품명|규격) 키를 만들고
    - 목록의 (품명|규격)와 정규화 비교해 일치/불일치를 기록.
    반환: summary(dict), map_df(DataFrame), mis_df(DataFrame)
    """
    import re
    import pandas as pd

    ul  = wb['일위대가']
    lst = wb['일위대가목록']

    # 일위대가 헤더 포지션
    ul_hr, ul_pos_full = find_cols(ul, {
        '품명','규격','단위','수량','합계 단가','합계금액',
        '재료비 단가','재료비 금액','노무비 단가','노무비 금액','경비 단가','경비 금액','비고'
    })
    COL = {
        "품명": ul_pos_full['품명'],
        "규격": ul_pos_full['규격'],
        "단위": ul_pos_full['단위'],
        "수량": ul_pos_full['수량'],
    }

    # 목록 필수 헤더
    ls_hr, ls_pos = find_cols(lst, {'코드','품명','규격'})

    # 목록 키
    def key_lst(r):
        return f"{norm(lst.cell(r, ls_pos['품명']).value)}|{norm(lst.cell(r, ls_pos['규격']).value)}"

    # 최근접 헤더 위로 탐색: 품명 존재 + (단위/수량 비어있음)
    def header_row_nearest(rr):
        for k in range(rr-1, ul_hr, -1):
            pname = str(ul.cell(k, COL["품명"]).value or "").strip()
            if not pname:
                continue
            unit = ul.cell(k, COL["단위"]).value
            qty  = ul.cell(k, COL["수량"]).value
            blank = {None, "", 0, "-", "—"}
            if unit in blank and qty in blank:
                return k
        return None

    # 헤더 키 생성: 규격이 비면 품명 내부의 '두 칸 이상 공백' 분리자를 이용해 앞 2토큰 → (품명|규격)
    def build_ul_header_key(r):
        pname = norm(ul.cell(r, COL['품명']).value)
        spec  = norm(ul.cell(r, COL['규격']).value)
        if not spec:
            toks = [t.strip() for t in re.split(r'[ \u3000]{2,}', pname) if t.strip()]
            if len(toks) >= 2:
                pname, spec = toks[0], toks[1]
        return f"{pname}|{spec}".strip("|")

    # '일위대가' 외부 참조 파서
    UL_CELLREF_RE = re.compile(r"(?:'?)일위대가(?:'?)!\$?([A-Z]{1,3})\$?(\d+)", re.I)

    mappings, records = [], []
    checked = 0

    for r in range(ls_hr + 1, lst.max_row + 1):
        list_key = key_lst(r)
        headers = set()
        fcount = 0

        # 행 전체의 수식을 스캔
        for c in range(1, lst.max_column + 1):
            cell = lst.cell(r, c)
            if cell.data_type != 'f' or not isinstance(cell.value, str):
                continue
            if "일위대가" not in cell.value:
                continue
            fcount += 1
            for col, row in UL_CELLREF_RE.findall(cell.value):
                rr = int(row)
                hdr = header_row_nearest(rr)
                if hdr:
                    headers.add(hdr)
                checked += 1

        hdr_row = sorted(list(headers))[0] if headers else None
        hdr_key = build_ul_header_key(hdr_row) if hdr_row else None

        mappings.append({
            "일위대가목록_행": r,
            "일위대가목록_품명|규격": list_key,
            "매핑_헤더행": hdr_row,
            "매핑_헤더_품명|규격": hdr_key,
            "참조셀_수": fcount
        })

        # 불일치 조건: 헤더키가 있고, 정규화 비교로 다르면
        if hdr_key and norm_key(list_key) != norm_key(hdr_key):
            records.append({
                "일위대가목록_행": r,
                "일위대가목록_품명|규격": list_key,
                "매핑_헤더행": hdr_row,
                "매핑_헤더_품명|규격": hdr_key
            })

        # 참조 전혀 없는데 금액·단가가 직접입력인 경우는 기존 로직에서 처리하던 케이스인데,
        # 여기선 목록 금액열 스캔 로직을 간소화했으므로, 필요시 원본 로직을 이어붙일 수 있음.

    import pandas as pd
    map_df = pd.DataFrame(mappings)
    mis_df = pd.DataFrame(records)
    summary = {
        "B_참조셀": int(checked),
        "B_매핑된_행": int(len(map_df)),
        "B_불일치": int(len(mis_df))
    }
    return summary, map_df, mis_df
def run_check_C(wb):
    """
    C (범용화): 공종별내역서의 각 행에서 '모든 외부 시트 참조'를 인식하여
    참조 대상 (시트, 행)의 (품명|규격) 키와 현재 행 키를 비교한다.
    외부 참조가 전혀 없고 '합계 단가' 셀이 수식이 아니며 0/공백이 아닌 경우 → 불일치(값 직접입력).
    반환: summary(dict), df(pandas.DataFrame)
    """
    import re, math
    import pandas as pd
    from collections import Counter

    src_name = '공종별내역서'
    if src_name not in wb.sheetnames:
        raise KeyError(f"Worksheet {src_name} does not exist.")

    wbs = wb[src_name]

    # 필요한 헤더 위치 (합계 단가는 반드시 필요)
    wbs_hr, wbs_pos = find_cols(wbs, {'품명','규격','합계 단가'})

    # (품명|규격) 키 함수
    def key_wbs(r):
        return f"{wbs.cell(r, wbs_pos['품명']).value}|{wbs.cell(r, wbs_pos['규격']).value}"

    # 범용 시트 참조 패턴: '시트명'!$A$1 또는 시트명!A1 등
    # 그룹: (quoted_sheet | bare_sheet) , col , row
    SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([^'!:]+))!\$?([A-Z]{1,3})\$?(\d+)", re.I)

    def get_key_from_sheet(sheet_name, rownum):
        """대상 시트에서 {'품명','규격'} 헤더가 발견되면 같은 행의 (품명|규격) 키를 반환. 없으면 None."""
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        try:
            hr, pos = find_cols(ws, {'품명','규격'})
        except Exception:
            return None
        r = int(rownum)
        if r <= hr or r > ws.max_row:
            return None
        return f"{ws.cell(r, pos['품명']).value}|{ws.cell(r, pos['규격']).value}"

    records = []
    rows_with_direct_refs = 0
    cnt_value_direct_input = 0

    for r in range(wbs_hr+1, wbs.max_row+1):
        wkey = key_wbs(r)
        if wkey == "None|None":
            continue

        # 행 내 모든 셀의 수식을 스캔하여 외부 시트 참조 수집(자기 시트 제외)
        refs = []
        for c in range(1, wbs.max_column+1):
            cell = wbs.cell(r, c)
            if cell.data_type == 'f' and isinstance(cell.value, str):
                for qsheet, bsheet, col, row in SHEET_REF_RE.findall(cell.value):
                    target_sheet = re.sub(r".*\(", "", (qsheet or bsheet or "").lstrip("=+").strip().strip("'"))
                    if not target_sheet or target_sheet == src_name:
                        continue
                    refs.append((target_sheet, int(row)))

        if refs:
            # 대표 참조 선택: (시트,행) 최빈값
            rows_with_direct_refs += 1
            cnt = Counter(refs)
            rep_sheet, rep_row = max(cnt.items(), key=lambda kv: kv[1])[0]
            tkey = get_key_from_sheet(rep_sheet, rep_row) or ""
            match_status = "일치" if norm_key(wkey) == norm_key(tkey) else "불일치"
            records.append({
                "행": r,
                "참조유형": f"{rep_sheet} 참조",
                "공종_키(품명|규격)": wkey,
                "참조_키(품명|규격)": tkey,
                "대표참조시트": rep_sheet,
                "대표참조행": rep_row,
                "일치여부": match_status
            })
            continue

        # 외부참조 없음 → '합계 단가' 직접입력 여부 체크
        cprice = wbs.cell(r, wbs_pos['합계 단가'])
        if cprice.data_type != 'f':
            val = cprice.value
            is_number = isinstance(val, (int, float)) and not isinstance(val, bool)
            nonzero = is_number and not (val == 0 or (isinstance(val, float) and math.isclose(val, 0.0, abs_tol=1e-9)))
            if nonzero:
                records.append({
                    "행": r, "참조유형": "값 직접입력",
                    "공종_키(품명|규격)": wkey,
                    "참조_키(품명|규격)": "",
                    "대표참조시트": "",
                    "대표참조행": "",
                    "일치여부": "불일치",
                    "입력값(합계단가)": val
                })
                cnt_value_direct_input += 1

    df = pd.DataFrame(records)
    summary = {
        "C_검사대상_행수(직접참조 보유)": int(rows_with_direct_refs),
        "C_일치": int((df["일치여부"]=="일치").sum()) if not df.empty else 0,
        "C_불일치": int((df["일치여부"]=="불일치").sum()) if not df.empty else 0,
        "C_값직접입력_불일치": int(cnt_value_direct_input)
    }
    return summary, df



#00000000000000000000 - D검사
def run_check_D(wb):
    """
    D (lenient): 공종별집계표의 재/노/경 단가 수식에서 외부 시트 참조를 수집하고,
    대표 참조(최빈 (시트, 행))의 (품명|규격) 키를 얻어 집계표 행 키와 비교한다.
    비교는 '품명(왼쪽)'만 수행한다. 헤더 추적은 최근접 헤더 기준.
    반환: summary(dict), map_df(DataFrame), mis_df(DataFrame)
    """
    import re
    import pandas as pd
    from collections import Counter

    s_sum = '공종별집계표'
    if s_sum not in wb.sheetnames:
        raise ValueError("공종별집계표 시트를 찾지 못했습니다.")
    ws_sum = wb[s_sum]

    # 집계표 필수 헤더
    sum_hr, sum_pos = find_cols(ws_sum, {'품명','규격','재료비 단가','노무비 단가','경비 단가'})

    # 키 생성(집계표)
    def key_sum(r):
        return f"{norm(ws_sum.cell(r, sum_pos['품명']).value)}|{norm(ws_sum.cell(r, sum_pos['규격']).value)}"

    # 범용 외부 참조 파서
    SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([^'!:]+))!\$?([A-Z]{1,3})\$?(\d+)", re.I)

    def left_name(k: str) -> str:
        name = (k.split("|",1)+[""])[0]
        name = re.sub(r"[，,]", " ", str(name))
        name = re.sub(r"\s+", " ", name.replace("\u3000"," ").strip())
        return name

    def pick_key_from_target(sheet_name, row_idx):
        if sheet_name not in wb.sheetnames:
            return None
        ws_t = wb[sheet_name]
        try:
            # 최근접 헤더 우선 (파일 상단에 정의된 key_from_header_or_same 사용)
            return key_from_header_or_same(ws_t, row_idx)
        except Exception:
            try:
                return key_from_same_row(ws_t, row_idx)
            except Exception:
                return None

    targets = [
        ('재료비 단가', sum_pos['재료비 단가']),
        ('노무비 단가', sum_pos['노무비 단가']),
        ('경비 단가',   sum_pos['경비 단가']),
    ]

    mappings, mismatches = [], []
    checked = 0

    for r in range(sum_hr + 1, ws_sum.max_row + 1):
        cur_key = key_sum(r)
        ref_pairs = []

        # 행의 세 단가 셀에서 외부 참조 수집
        for label, c in targets:
            cell = ws_sum.cell(r, c)
            val = cell.value
            if not isinstance(val, str) or "!" not in val:
                continue
            for m in SHEET_REF_RE.finditer(val):
                qsheet = m.group(1) or m.group(2) or ""
                rownum = int(m.group(4))
                tgt_sheet = qsheet.strip().strip("'").lstrip("=")
                if tgt_sheet and tgt_sheet != s_sum:
                    ref_pairs.append((tgt_sheet, rownum))
                    checked += 1

        rep_sheet, rep_row, rep_key = None, None, None
        if ref_pairs:
            (rep_sheet, rep_row), _ = Counter(ref_pairs).most_common(1)[0]
            rep_key = pick_key_from_target(rep_sheet, rep_row)

        mappings.append({
            "집계표_행": r,
            "집계표_품명|규격": cur_key,
            "대표참조_시트": rep_sheet,
            "대표참조_행": rep_row,
            "대표참조_품명|규격": rep_key,
            "참조개수": len(ref_pairs),
        })

        # ★ lenient 비교: '품명'만 비교
        if rep_key and left_name(cur_key) != left_name(rep_key):
            mismatches.append({
                "집계표_행": r,
                "집계표_품명|규격": cur_key,
                "대표참조_시트": rep_sheet,
                "대표참조_행": rep_row,
                "대표참조_품명|규격": rep_key,
            })

    map_df = pd.DataFrame(mappings)
    mis_df = pd.DataFrame(mismatches)
    summary = {
        "D_참조셀": int(checked),
        "D_매핑된_행": int(len(map_df)),
        "D_불일치": int(len(mis_df)),
    }
    return summary, map_df, mis_df
def key_from_header_or_same(ws, r):
    """내역서형이면 '최근접 헤더행'을 직접 찾아 (품명|규격) 키를 만들고,
    실패하면 같은 행 키. 내역서형 아니면 같은 행 키."""
    if is_like_detail_sheet(ws):
        try:
            det_hr, det_pos = find_cols(ws, {'품명','규격','단위','수량'})
        except Exception:
            return key_from_same_row(ws, r)
        def norm_s(x):
            if x is None: return ""
            return str(x).replace("\u3000"," ").strip()
        def is_name_header_row(rr):
            pname = norm_s(ws.cell(rr, det_pos['품명']).value)
            if not pname:
                return False
            unit  = ws.cell(rr, det_pos['단위']).value
            qty   = ws.cell(rr, det_pos['수량']).value
            blank = {None, "", 0, "-", "—"}
            if unit not in blank or qty not in blank:
                return False
            # '합계' 텍스트 자체는 배제
            s = pname.replace(" ", "").replace("\u3000","")
            if "합계" in s:
                return False
            return True
        # 1) 위로 스캔하여 가장 가까운 헤더
        for k in range(r-1, det_hr, -1):
            if is_name_header_row(k):
                kk = key_from_same_row(ws, k)
                if kk: return kk
        # 2) 폴백: 위로 품명 텍스트가 있는 첫 행
        for k in range(r-1, det_hr, -1):
            pname = norm_s(ws.cell(k, det_pos['품명']).value)
            if pname:
                kk = key_from_same_row(ws, k)
                if kk: return kk
    # 최종 폴백: 같은 행
    return key_from_same_row(ws, r)

def run_check_E(wb):
    """
    E: 단가대비표의 ❶재료비 적용단가, ❷노무비 셀에서 '외부 시트 참조'를 찾아
       단가대비표 (품명|규격) ↔ 참조 시트 키를 비교한다.
       - 장비 단가산출서 참조: (품명|사양)
       - 그 외 시트: (품명|규격)
       - 필요 시 품평/사양을 fallback 으로 시도

    반환: summary(dict), df(DataFrame)
    """
    import re
    import pandas as pd

    s_dv = '단가대비표'
    if s_dv not in wb.sheetnames:
        raise KeyError(f"Worksheet {s_dv} does not exist.")
    dv = wb[s_dv]

    # 헤더 포지션
    dv_hr, dv_pos = find_cols(dv, {'품명','규격','재료비 적용단가','노무비'})
    p_col, g_col  = dv_pos['품명'], dv_pos['규격']
    c_cols = [(dv_pos['재료비 적용단가'], '재료비 적용단가'),
              (dv_pos['노무비'],        '노무비')]

    # 외부시트 참조 파서 (C/D와 동일 철학)
    SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([^'!:]+))!\$?([A-Z]{1,3})\$?(\d+)", re.I)
    def clean_sheet_name(sn: str) -> str:
        if not sn: return ""
        s = str(sn).lstrip("=+").strip().strip("'")
        s = re.sub(r".*\(", "", s)   # 예: TRUNC(시트!A1) 보호
        return s

    def key_from_dynamic(sheet_name: str, rownum: int):
        """
        장비 단가산출서 → (품명|사양) 우선
        일반 시트      → (품명|규격) 우선
        Fallback       → (품명|사양), (품평|규격)
        """
        if sheet_name not in wb.sheetnames:
            return None, None
        ws = wb[sheet_name]
        r = int(rownum)
        title = re.sub(r"\s+", "", sheet_name)

        if ("장비" in title and "단가산출서" in title):
            orders = [('품명','사양'), ('품명','규격'), ('품평','규격')]
        else:
            orders = [('품명','규격'), ('품명','사양'), ('품평','규격')]

        for h1, h2 in orders:
            try:
                hr, pos = find_cols(ws, {h1, h2})
            except Exception:
                continue
            if r <= hr or r > ws.max_row:
                continue
            key = f"{norm(ws.cell(r, pos[h1]).value)}|{norm(ws.cell(r, pos[h2]).value)}"
            return key, f"{h1}|{h2}"
        return None, None

    records = []
    for r in range(dv_hr+1, dv.max_row+1):
        base_key = f"{norm(dv.cell(r, p_col).value)}|{norm(dv.cell(r, g_col).value)}"
        if base_key == "None|None":
            continue

        for cidx, label in c_cols:
            cell = dv.cell(r, cidx)
            if cell.data_type != 'f' or not isinstance(cell.value, str):
                continue
            fml = cell.value

            # 시트 참조 수집 (자기 시트 제외)
            refs = []
            for q,b,col,row in SHEET_REF_RE.findall(fml):
                ts = clean_sheet_name(q or b)
                if not ts or ts == s_dv:
                    continue
                if ts in wb.sheetnames:
                    refs.append((ts, int(row)))
            if not refs:
                continue

            ts, rr = refs[0]                # 대표: 첫 참조
            ref_key, used_hdr = key_from_dynamic(ts, rr)
            match_status = None
            if ref_key is not None:
                match_status = "일치" if norm_key(base_key) == norm_key(ref_key) else "불일치"

            records.append({
                "행": r,
                "대상열": label,
                "수식": fml,
                "참조시트": ts,
                "참조행": rr,
                "단가대비표_키(품명|규격)": base_key,
                "참조_키": ref_key,
                "참조키_사용헤더": used_hdr,
                "참조셀_수": len(refs),
                "일치여부": match_status
            })

    df = pd.DataFrame(records)
    summary = {
        "E_총검사셀": len(df),
        "E_일치": int((df["일치여부"]=="일치").sum()) if not df.empty else 0,
        "E_불일치": int((df["일치여부"]=="불일치").sum()) if not df.empty else 0,
        "E_참조키_None": int(df["참조_키"].isna().sum()) if not df.empty else 0,
    }
    return summary, df

# ------------ main ------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--infile", required=True)
    ap.add_argument("--outdir", required=True)
    args = ap.parse_args()

    # 원본 파일명 꼬리 추출 + 파일명 생성
    src_stem = os.path.splitext(os.path.basename(args.infile))[0]
    if src_stem.startswith("(") and ")" in src_stem:
        src_tail = src_stem.split(")", 1)[1].lstrip("_")
    else:
        src_tail = src_stem

    def mk(outdir, base, ext):
        return os.path.join(outdir, f"{base}_{src_tail}.{ext}")


    outdir_final = os.path.join(args.outdir, f"{src_tail} 검사결과")
    os.makedirs(outdir_final, exist_ok=True)

    wb = load_workbook(args.infile, data_only=False, keep_vba=True)

    sumA, dfA = run_check_A(wb)
    sumB, mapB, misB = run_check_B(wb)
    sumC, dfC = run_check_C(wb)
    sumD, mapD, misD = run_check_D(wb)
    sumE, dfE = run_check_E(wb)

    # 사용 예시
    infile = "/mnt/data/(전처리완료)토목 내역서_20250912.xlsx"
    outdir = "/mnt/data/검사_결과"
    
    # A
    a_full_xlsx  = mk(outdir_final, "일위대가 검사_전체", "xlsx")

    dfA.to_excel(a_full_xlsx, index=False)

    dfA[dfA["일치여부"]=="일치"].to_excel(mk(outdir_final, "일위대가 검사_일치", "xlsx"), index=False)
  
    dfA[dfA["일치여부"]=="불일치"].to_excel(mk(outdir_final, "일위대가 검사_불일치", "xlsx"), index=False)
   

    # B
    outB_map_xlsx = mk(outdir_final, "일위대가 목록_전체", "xlsx")
   
    outB_mis_xlsx = mk(outdir_final, "일위대가 목록_불일치", "xlsx")
 
    mapB.to_excel(outB_map_xlsx, index=False)

    misB.to_excel(outB_mis_xlsx, index=False)


    # C
    c_full_xlsx  = mk(outdir_final, "공종별 내역서_전체", "xlsx")

    dfC.to_excel(c_full_xlsx, index=False)
 
    dfC[dfC["일치여부"]=="일치"].to_excel(mk(outdir_final, "공종별 내역서_일치", "xlsx"), index=False)
    
    dfC[dfC["일치여부"]=="불일치"].to_excel(mk(outdir_final, "공종별 내역서_불일치", "xlsx"), index=False)
    
    #  D 
    outD_map_xlsx = mk(outdir_final, "공종별 집계표_전체", "xlsx")
    outD_mis_xlsx = mk(outdir_final, "공종별 집계표_불일치", "xlsx")

    safe_to_excel(mapD, outD_map_xlsx, index=False)
    safe_to_excel(misD, outD_mis_xlsx, index=False)

    
    # E 저장: 전체 / 불일치
    e_all_xlsx = mk(outdir_final, "단가대비표 검사(건축)_전체", "xlsx")
    e_mis_xlsx = mk(outdir_final, "단가대비표 검사(건축)_불일치", "xlsx")
    safe_to_excel(dfE, e_all_xlsx, index=False)

    # ← 여기만 가드 추가
    if "일치여부" in dfE.columns:
        safe_to_excel(dfE[dfE["일치여부"] == "불일치"], e_mis_xlsx, index=False)
    else:
        # 결과 0건인 파일도 '빈 표'로 저장되도록
        safe_to_excel(dfE.head(0), e_mis_xlsx, index=False)


    # ▼ 요약 출력 (C 출력 저장 직후에 추가)
    summary = {**sumA, **sumB, **sumC}
    summary.update(sumD)
    summary.update(sumE) 

    print("\n검사 완료. 요약:")
    print(f" 일위대가 검사: 참조 {summary.get('A_검사한_참조', 0)}건, 일치 {summary.get('A_일치', 0)}, 불일치 {summary.get('A_불일치', 0)}")
    print(f" 일위대가 목록 검사: 참조셀 {summary.get('B_참조셀', 0)}, 헤더매핑된 행 {summary.get('B_매핑된_행', 0)}, 불일치 {summary.get('B_불일치', 0)}")
    print(f" 공종별 내역서 검사: 검사대상 {summary.get('C_검사대상_행수(직접참조 보유)', 0)}행, 일치 {summary.get('C_일치', 0)}, 불일치 {summary.get('C_불일치', 0)}")
    print(f" 공종별 집계표 검사: 참조셀 {summary.get('D_참조셀', 0)}, 헤더매핑된 행 {summary.get('D_매핑된_행', 0)}, 불일치 {summary.get('D_불일치', 0)}")
    print(f" 단가대비표 검사(건축): 검사셀 {summary.get('E_총검사셀', 0)}, 일치 {summary.get('E_일치', 0)}, 불일치 {summary.get('E_불일치', 0)}")


    print("outdir:", args.outdir)

if __name__ == "__main__":
    sys.exit(main())

# === (추가) 간이 웹서버 모드 ===
# 사용:  python Matchflag_rev2_9.py --serve
# 또는:  WEB=1 python Matchflag_rev2_9.py
try:
    import io, tempfile, shutil, json
    from flask import Flask, request, send_from_directory, jsonify
except Exception:
    # 웹 모드가 필요 없는 환경이면 무시
    pass

def _run_matchflag_in_memory(xlsx_fp, workdir):
    """
    업로드 받은 엑셀 파일 핸들(xlsx_fp)을 임시 경로에 저장하고 기존 로직을 호출.
    outdir 아래 생성된 결과 파일 경로들을 리턴.
    """
    import os
    from openpyxl import load_workbook

    os.makedirs(workdir, exist_ok=True)
    src_path = os.path.join(workdir, "input.xlsx")
    with open(src_path, "wb") as f:
        f.write(xlsx_fp.read())

    # outdir = workdir/results
    outdir = os.path.join(workdir, "results")
    os.makedirs(outdir, exist_ok=True)

    # 원본 main 로직을 함수형으로 재사용
    # 여기서는 main() 내부를 간접 실행: workbook을 직접 열어 하위 run_check_* 호출
    wb = load_workbook(src_path, data_only=False, keep_vba=True)

    # ===== 아래는 기존 main()과 동일한 저장 로직 재현 =====
    src_stem = os.path.splitext(os.path.basename(src_path))[0]
    src_tail = src_stem

    def mk(outdir_final, base, ext):
        return os.path.join(outdir_final, f"{base}_{src_tail}.{ext}")

    outdir_final = os.path.join(outdir, f"{src_tail} 검사결과")
    os.makedirs(outdir_final, exist_ok=True)

    sumA, dfA = run_check_A(wb)
    sumB, mapB, misB = run_check_B(wb)
    sumC, dfC = run_check_C(wb)
    sumD, mapD, misD = run_check_D(wb)
    sumE, dfE = run_check_E(wb)

    paths = []

    a_full_xlsx  = mk(outdir_final, "일위대가 검사_전체", "xlsx")
    dfA.to_excel(a_full_xlsx, index=False); paths.append(a_full_xlsx)
    p = mk(outdir_final, "일위대가 검사_일치", "xlsx")
    dfA[dfA["일치여부"]=="일치"].to_excel(p, index=False); paths.append(p)
    p = mk(outdir_final, "일위대가 검사_불일치", "xlsx")
    dfA[dfA["일치여부"]=="불일치"].to_excel(p, index=False); paths.append(p)

    outB_map_xlsx = mk(outdir_final, "일위대가 목록_전체", "xlsx")
    outB_mis_xlsx = mk(outdir_final, "일위대가 목록_불일치", "xlsx")
    mapB.to_excel(outB_map_xlsx, index=False); paths.append(outB_map_xlsx)
    misB.to_excel(outB_mis_xlsx, index=False); paths.append(outB_mis_xlsx)

    c_full_xlsx  = mk(outdir_final, "공종별 내역서_전체", "xlsx")
    dfC.to_excel(c_full_xlsx, index=False); paths.append(c_full_xlsx)
    p = mk(outdir_final, "공종별 내역서_일치", "xlsx")
    dfC[dfC["일치여부"]=="일치"].to_excel(p, index=False); paths.append(p)
    p = mk(outdir_final, "공종별 내역서_불일치", "xlsx")
    dfC[dfC["일치여부"]=="불일치"].to_excel(p, index=False); paths.append(p)

    outD_map_xlsx = mk(outdir_final, "공종별 집계표_전체", "xlsx")
    outD_mis_xlsx = mk(outdir_final, "공종별 집계표_불일치", "xlsx")
    safe_to_excel(mapD, outD_map_xlsx, index=False); paths.append(outD_map_xlsx)
    safe_to_excel(misD, outD_mis_xlsx, index=False); paths.append(outD_mis_xlsx)

    e_all_xlsx = mk(outdir_final, "단가대비표 검사(건축)_전체", "xlsx")
    e_mis_xlsx = mk(outdir_final, "단가대비표 검사(건축)_불일치", "xlsx")
    safe_to_excel(dfE, e_all_xlsx, index=False); paths.append(e_all_xlsx)
    if "일치여부" in dfE.columns:
        safe_to_excel(dfE[dfE["일치여부"] == "불일치"], e_mis_xlsx, index=False)
    else:
        safe_to_excel(dfE.head(0), e_mis_xlsx, index=False)
    paths.append(e_mis_xlsx)

    # 요약 로그
    summary = {**sumA, **sumB, **sumC}; summary.update(sumD); summary.update(sumE)
    log = []
    log.append("검사 완료. 요약:")
    log.append(f" 일위대가 검사: 참조 {summary.get('A_검사한_참조', 0)}건, 일치 {summary.get('A_일치', 0)}, 불일치 {summary.get('A_불일치', 0)}")
    log.append(f" 일위대가 목록 검사: 참조셀 {summary.get('B_참조셀', 0)}, 헤더매핑된 행 {summary.get('B_매핑된_행', 0)}, 불일치 {summary.get('B_불일치', 0)}")
    log.append(f" 공종별 내역서 검사: 검사대상 {summary.get('C_검사대상_행수(직접참조 보유)', 0)}행, 일치 {summary.get('C_일치', 0)}, 불일치 {summary.get('C_불일치', 0)}")
    log.append(f" 공종별 집계표 검사: 참조셀 {summary.get('D_참조셀', 0)}, 헤더매핑된 행 {summary.get('D_매핑된_행', 0)}, 불일치 {summary.get('D_불일치', 0)}")
    log.append(f" 단가대비표 검사(건축): 검사셀 {summary.get('E_총검사셀', 0)}, 일치 {summary.get('E_일치', 0)}, 불일치 {summary.get('E_불일치', 0)}")
    return log, paths, outdir

def _as_downloads(base_dir, paths):
    files = []
    for p in paths:
        name = os.path.basename(p)
        url  = f"/download/{name}"
        files.append({"name": name, "url": url})
    return files

def _create_app_for_matchflag():
    app = Flask(__name__)
    # 결과 파일 제공용 폴더
    DL_DIR = tempfile.mkdtemp(prefix="mf_out_")

    @app.post("/api/matchflag")
    def api_matchflag():
        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "log": "파일이 없습니다."}), 400
        # 작업 디렉토리
        work = tempfile.mkdtemp(prefix="mf_work_")
        log, paths, outdir = _run_matchflag_in_memory(f.stream, workdir=work)
        # 결과물 복사 → 고정 다운로드 디렉토리
        copied = []
        for p in paths:
            dst = os.path.join(DL_DIR, os.path.basename(p))
            shutil.copy2(p, dst)
            copied.append(dst)
        return jsonify({"ok": True, "log": "\n".join(log), "files": _as_downloads(DL_DIR, copied)})

    @app.get("/download/<path:fname>")
    def download_file(fname):
        return send_from_directory(DL_DIR, fname, as_attachment=True)

    return app

if __name__ == "__main__":
    import os, sys
    if "--serve" in sys.argv or os.environ.get("WEB") == "1":
        app = _create_app_for_matchflag()
        app.run(host="0.0.0.0", port=8000, debug=False)
